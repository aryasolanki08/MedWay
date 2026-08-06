from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from django.http import HttpResponse
from accounts.permissions import PharmacyScopeMixin, HasPharmacy
from billing.models import Bill
from billing.serializers import BillSerializer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class BillViewSet(PharmacyScopeMixin, viewsets.ModelViewSet):
    queryset = Bill.objects.all().order_by('-created_at')
    serializer_class = BillSerializer
    permission_classes = [HasPharmacy]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        # Precalculate sequence map for the current pharmacy
        user = request.user
        pharmacy = None
        if not user.is_superuser:
            try:
                pharmacy = user.staff_profile.pharmacy
            except Exception:
                pass
                
        if pharmacy:
            all_ids = list(Bill.objects.filter(pharmacy=pharmacy).order_by('id').values_list('id', flat=True))
            seq_map = {b_id: idx + 1 for idx, b_id in enumerate(all_ids)}
        else:
            all_bills = Bill.objects.order_by('pharmacy_id', 'id').values_list('id', 'pharmacy_id')
            seq_map = {}
            pharmacy_counts = {}
            for b_id, ph_id in all_bills:
                pharmacy_counts[ph_id] = pharmacy_counts.get(ph_id, 0) + 1
                seq_map[b_id] = pharmacy_counts[ph_id]
                
        page = self.paginate_queryset(queryset)
        serializer_context = self.get_serializer_context()
        serializer_context['seq_map'] = seq_map
        
        if page is not None:
            serializer = self.get_serializer(page, many=True, context=serializer_context)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True, context=serializer_context)
        from rest_framework.response import Response
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.get_queryset()
        
        # Search query filtering
        search_query = request.query_params.get('search', '')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(customer_name__icontains=search_query) | 
                Q(id__icontains=search_query)
            )
            
        # Date filtering
        start_date = request.query_params.get('start_date', '')
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
            
        # End date filtering
        end_date = request.query_params.get('end_date', '')
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)

        # Precalculate sequence map for Excel export
        user = request.user
        pharmacy = None
        if not user.is_superuser:
            try:
                pharmacy = user.staff_profile.pharmacy
            except Exception:
                pass
                
        if pharmacy:
            all_ids = list(Bill.objects.filter(pharmacy=pharmacy).order_by('id').values_list('id', flat=True))
            seq_map = {b_id: idx + 1 for idx, b_id in enumerate(all_ids)}
        else:
            all_bills = Bill.objects.order_by('pharmacy_id', 'id').values_list('id', 'pharmacy_id')
            seq_map = {}
            pharmacy_counts = {}
            for b_id, ph_id in all_bills:
                pharmacy_counts[ph_id] = pharmacy_counts.get(ph_id, 0) + 1
                seq_map[b_id] = pharmacy_counts[ph_id]

        # Generate workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: Sales Summary
        ws1 = wb.active
        ws1.title = "Sales Summary"
        
        headers1 = [
            "Invoice ID", "Date", "Customer Name", "Customer Phone", 
            "Items Count", "Discount", "Total Amount", "Handled By"
        ]
        for col_idx, h in enumerate(headers1, 1):
            ws1.cell(row=3, column=col_idx, value=h)
            
        for row_idx, bill in enumerate(queryset, 4):
            items_count = bill.items.count()
            created_by_name = bill.created_by.user.username if (bill.created_by and bill.created_by.user) else "System/Admin"
            seq_id = seq_map.get(bill.id, bill.id)
            row_data = [
                f"#{seq_id}",
                bill.created_at.strftime("%Y-%m-%d %H:%M:%S") if bill.created_at else "",
                bill.customer_name or "Walk-in Customer",
                bill.customer_phone or "N/A",
                items_count,
                float(bill.discount),
                float(bill.total_amount),
                created_by_name
            ]
            for col_idx, val in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=val)
            
        # Sheet 2: Detailed Line Items
        ws2 = wb.create_sheet(title="Sales Items Detail")
        
        headers2 = [
            "Invoice ID", "Customer Name", "Medicine Name", "Category", 
            "Batch Number", "Expiry Date", "Quantity", "Unit Price", "Subtotal"
        ]
        for col_idx, h in enumerate(headers2, 1):
            ws2.cell(row=3, column=col_idx, value=h)
            
        current_row = 4
        for bill in queryset:
            seq_id = seq_map.get(bill.id, bill.id)
            for item in bill.items.all():
                row_data = [
                    f"#{seq_id}",
                    bill.customer_name or "Walk-in Customer",
                    item.medicine.name if item.medicine else "Unknown",
                    item.medicine.category if item.medicine else "",
                    item.medicine.batch_number if item.medicine else "",
                    item.medicine.expiry_date.strftime("%Y-%m-%d") if (item.medicine and item.medicine.expiry_date) else "",
                    item.quantity,
                    float(item.unit_price),
                    float(item.subtotal)
                ]
                for col_idx, val in enumerate(row_data, 1):
                    ws2.cell(row=current_row, column=col_idx, value=val)
                current_row += 1
                
        # Helper function to style sheets
        def style_ws(ws, title, number_formats):
            ws.views.sheetView[0].showGridLines = True
            
            # Title block in row 1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = title
            title_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40
            
            # Headers style in row 3
            header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
            header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            ws.row_dimensions[3].height = 25
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=3, column=col)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = thin_border
                
            # Data row styling
            for r in range(4, ws.max_row + 1):
                ws.row_dimensions[r].height = 20
                row_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid') if r % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=r, column=col)
                    c.fill = row_fill
                    c.border = thin_border
                    c.font = Font(name='Arial', size=10)
                    
                    header_val = ws.cell(row=3, column=col).value
                    if header_val in number_formats:
                        c.number_format = number_formats[header_val]
                        
                    if isinstance(c.value, (int, float)):
                        c.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        c.alignment = Alignment(horizontal='left', vertical='center')
                        
            # Auto width
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1:
                        continue
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        style_ws(ws1, "Sales History Summary", {
            "Discount": '"₹"#,##0.00',
            "Total Amount": '"₹"#,##0.00',
            "Items Count": '#,##0'
        })
        
        style_ws(ws2, "Sales Line Items Detail", {
            "Quantity": '#,##0',
            "Unit Price": '"₹"#,##0.00',
            "Subtotal": '"₹"#,##0.00'
        })
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_history.xlsx"'
        wb.save(response)
        return response
